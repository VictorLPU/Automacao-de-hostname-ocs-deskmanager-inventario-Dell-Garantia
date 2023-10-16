import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from desk import main_desk
from ocs import main_ocs
from dell import main_dell
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from selenium.webdriver.chrome.options import Options


def dados_excel(caminho_arquivo):
    meses = {'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
             'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
             'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'}

    arquivo_excel = openpyxl.load_workbook(caminho_arquivo)
    aba = arquivo_excel['RJ']

    for indice_linha, linha in enumerate(aba.iter_rows(values_only=True)):

        if linha[4] is not None or linha[5] is not None or linha[6] is not None or \
                linha[7] is not None or linha[8] is not None:
            continue
        for indice, valor in enumerate(linha):
            if valor is not None:
                if indice_linha != 0:
                    if indice == 0:
                        while True:
                            try:
                                serial_number = main_ocs(valor)
                                aba[f"J{indice_linha + 1}"] = f'{serial_number}'
                                if serial_number == 'No matching records found':
                                    aba[f"F{indice_linha + 1}"] = f'{serial_number}'
                                    break

                                aba[f"F{indice_linha + 1}"] = f'Ok'
                                break
                            except:
                                pass

                        if serial_number == 'No matching records found':
                            continue

                        while True:
                            try:
                                try:
                                    data_inicio, data_vencimento = main_dell(serial_number)
                                    continuar = False
                                    break
                                except:
                                    continuar = True
                                    aba[f"E{indice_linha + 1}"] = f'Dell não encontrado'
                                    aba[f"F{indice_linha + 1}"] = f'Dell não encontrado'
                                    arquivo_excel.save(caminho_arquivo)
                                    break
                            except:
                                pass
                        if continuar:
                            continue
                        data_inicio = data_inicio.split(' ')
                        data_inicio[1] = meses[data_inicio[1]]
                        # Verificação mais de 5 anos
                        verificacao_data = datetime.date(int(data_inicio[2]), int(data_inicio[1]), int(data_inicio[0]))
                        data_atual = datetime.date.today()

                        # Subtraia 5 anos da data atual
                        data_limite = data_atual - datetime.timedelta(days=5 * 365)

                        # Verifique se a data desejada está há mais de 5 anos no passado
                        if verificacao_data >= data_limite:
                            aba[f"I{indice_linha + 1}"] = f'Menos de 5'
                        else:
                            aba[f"I{indice_linha + 1}"] = f'Mais de 5'
                        data_inicio = '-'.join(data_inicio)

                        data_vencimento = data_vencimento.split(' ')
                        data_vencimento[1] = meses[data_vencimento[1]]
                        data_vencimento = '-'.join(data_vencimento)
                        aba[f"G{indice_linha + 1}"] = f'{data_inicio}'
                        aba[f"H{indice_linha + 1}"] = f'{data_vencimento}'
                        while True:
                            try:
                                desk = main_desk(valor, data_inicio)
                                aba[f"E{indice_linha + 1}"] = f'{desk}'
                                arquivo_excel.save(caminho_arquivo)
                                print('foi')
                                break
                            except:
                                pass

    arquivo_excel.save(caminho_arquivo)

    aba = arquivo_excel['SP']

    for indice_linha, linha in enumerate(aba.iter_rows(values_only=True)):

        if linha[4] is not None or linha[5] is not None or linha[6] is not None or\
           linha[7] is not None or linha[8] is not None:
            continue
        for indice, valor in enumerate(linha):
            if valor is not None:
                if indice_linha != 0:
                    if indice == 0:
                        while True:
                            try:
                                serial_number = main_ocs(valor)
                                aba[f"J{indice_linha + 1}"] = f'{serial_number}'
                                if serial_number == 'No matching records found':
                                    break

                                aba[f"F{indice_linha + 1}"] = f'Ok'
                                break
                            except:
                                pass
                        if serial_number == 'No matching records found':
                            continue

                        while True:
                            try:
                                try:
                                    data_inicio, data_vencimento = main_dell(serial_number)
                                    continuar = False
                                    break
                                except:
                                    continuar = True
                                    aba[f"E{indice_linha + 1}"] = f'Dell não encontrado'
                                    aba[f"F{indice_linha + 1}"] = f'Dell não encontrado'
                                    arquivo_excel.save(caminho_arquivo)
                                    break
                            except:
                                pass
                        if continuar:
                            continue
                        data_inicio = data_inicio.split(' ')
                        data_inicio[1] = meses[data_inicio[1]]
                        # Verificação mais de 5 anos
                        verificacao_data = datetime.date(int(data_inicio[2]), int(data_inicio[1]), int(data_inicio[0]))
                        data_atual = datetime.date.today()

                        # Subtraia 5 anos da data atual
                        data_limite = data_atual - datetime.timedelta(days=5 * 365)

                        # Verifique se a data desejada está há mais de 5 anos no passado
                        if verificacao_data >= data_limite:
                            aba[f"I{indice_linha + 1}"] = f'Menos de 5'
                        else:
                            aba[f"I{indice_linha + 1}"] = f'Mais de 5'
                        data_inicio = '-'.join(data_inicio)

                        data_vencimento = data_vencimento.split(' ')
                        data_vencimento[1] = meses[data_vencimento[1]]
                        data_vencimento = '-'.join(data_vencimento)
                        aba[f"G{indice_linha + 1}"] = f'{data_inicio}'
                        aba[f"H{indice_linha + 1}"] = f'{data_vencimento}'
                        while True:
                            try:
                                desk = main_desk(valor, data_inicio)
                                aba[f"E{indice_linha + 1}"] = f'{desk}'
                                arquivo_excel.save(caminho_arquivo)
                                print('foi')
                                break
                            except:
                                pass

    arquivo_excel.save(caminho_arquivo)




if __name__ == '__main__':
    dados_excel('PlanilhasRJESP.xlsx')
